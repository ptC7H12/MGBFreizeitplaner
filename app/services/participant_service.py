"""Participant Service - Business-Logik für Teilnehmer"""
import logging
from io import BytesIO
from datetime import date
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models import Participant, Role, Event, Family, Ruleset
from app.services.price_calculator import PriceCalculator
from app.services.excel_service import ExcelService

logger = logging.getLogger(__name__)


class ParticipantService:
    """Service für Teilnehmer-Business-Logik"""

    @staticmethod
    def calculate_price_for_participant(
        db: Session,
        event_id: int,
        role_id: int,
        birth_date: date,
        family_id: Optional[int] = None
    ) -> float:
        """
        Berechnet den Preis für einen Teilnehmer.

        Args:
            db: Datenbank-Session
            event_id: ID des Events
            role_id: ID der Rolle
            birth_date: Geburtsdatum des Teilnehmers
            family_id: Optional - ID der Familie für Familienrabatte

        Returns:
            Berechneter Preis in Euro
        """
        # Event und Ruleset laden
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise ValueError(f"Event mit ID {event_id} nicht gefunden")

        ruleset = db.query(Ruleset).filter(Ruleset.id == event.ruleset_id).first()
        if not ruleset:
            raise ValueError(f"Ruleset für Event {event_id} nicht gefunden")

        # Rolle laden
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            raise ValueError(f"Rolle mit ID {role_id} nicht gefunden")

        # Alter berechnen
        age = event.start_date.year - birth_date.year
        if (event.start_date.month, event.start_date.day) < (birth_date.month, birth_date.day):
            age -= 1

        # Position in der Familie bestimmen (für Familienrabatte)
        family_children_count = 1
        if family_id:
            family = db.query(Family).filter(Family.id == family_id).first()
            if family:
                # Anzahl aktiver Kinder in der Familie
                children_in_family = db.query(Participant).filter(
                    Participant.family_id == family_id,
                    Participant.event_id == event_id,
                    Participant.is_active == True
                ).order_by(Participant.birth_date.desc()).all()

                # Position des neuen Kindes wäre am Ende
                family_children_count = len(children_in_family) + 1

        # Preis berechnen
        return PriceCalculator.calculate_participant_price(
            age=age,
            role_name=role.name,
            ruleset_data=ruleset.rules,
            family_children_count=family_children_count
        )

    @staticmethod
    def export_to_excel(participants: List[Participant], event: Event, db: Session) -> BytesIO:
        """
        Exportiert Teilnehmer als Excel-Datei.

        Args:
            participants: Liste der Teilnehmer
            event: Event-Objekt
            db: Datenbank-Session für Familienzahlungs-Berechnung

        Returns:
            BytesIO-Stream mit Excel-Datei
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Teilnehmerliste"

        # Header definieren
        headers = [
            "Nachname", "Vorname", "Geburtsdatum", "Alter", "Geschlecht",
            "Rolle", "Familie", "E-Mail", "Telefon", "Preis (€)",
            "Bezahlt (€)", "Offen (€)", "Adresse"
        ]

        # Spaltenbreiten setzen
        column_widths = {
            1: 20,  # Nachname
            2: 20,  # Vorname
            3: 15,  # Geburtsdatum
            4: 8,   # Alter
            5: 12,  # Geschlecht
            6: 15,  # Rolle
            7: 20,  # Familie
            8: 30,  # E-Mail
            9: 15,  # Telefon
            10: 12, # Preis
            11: 12, # Bezahlt
            12: 12, # Offen
            13: 40  # Adresse
        }

        # Header-Formatierung mit ExcelService
        ExcelService.apply_header_row(ws, headers, column_widths)

        # Daten schreiben
        for row_num, participant in enumerate(participants, 2):
            # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
            direct_payments = float(sum((payment.amount for payment in participant.payments), 0))

            # Anteilige Familienzahlungen berechnen
            family_payment_share = 0.0
            if participant.family_id:
                # Lade Familie mit Zahlungen und Teilnehmern
                from sqlalchemy.orm import joinedload
                family = db.query(Family).filter(Family.id == participant.family_id).options(
                    joinedload(Family.payments),
                    joinedload(Family.participants).joinedload(Participant.payments)
                ).first()

                if family:
                    # Gesamtzahlungen der Familie
                    family_total_payments = float(sum((payment.amount for payment in family.payments), 0))

                    # Berechne für jedes Familienmitglied die direkten Zahlungen und offenen Beträge
                    family_members_outstanding = []
                    total_outstanding = 0.0

                    for member in family.participants:
                        member_direct_payments = float(sum((payment.amount for payment in member.payments), 0))
                        member_outstanding = max(0.0, float(member.final_price) - member_direct_payments)
                        family_members_outstanding.append({
                            'participant_id': member.id,
                            'outstanding': member_outstanding
                        })
                        total_outstanding += member_outstanding

                    # Verteile die Familienzahlung proportional auf die offenen Beträge
                    if total_outstanding > 0:
                        # Finde den aktuellen Teilnehmer in der Liste
                        for member_data in family_members_outstanding:
                            if member_data['participant_id'] == participant.id:
                                # Anteilige Verteilung basierend auf offenen Beträgen
                                family_payment_share = (member_data['outstanding'] / total_outstanding) * family_total_payments
                                break
                    elif family_total_payments > 0:
                        # Wenn alle Beträge bezahlt sind, aber noch Familienzahlungen da sind,
                        # verteile proportional nach Sollpreis
                        family_total_price = float(sum((p.final_price for p in family.participants), 0))
                        if family_total_price > 0:
                            family_payment_share = (float(participant.final_price) / family_total_price) * family_total_payments

            total_paid = direct_payments + family_payment_share
            outstanding = float(participant.final_price) - total_paid

            ws.cell(row=row_num, column=1, value=participant.last_name)
            ws.cell(row=row_num, column=2, value=participant.first_name)
            ws.cell(row=row_num, column=3, value=participant.birth_date.strftime("%d.%m.%Y") if participant.birth_date else "")
            ws.cell(row=row_num, column=4, value=participant.age_at_event or "")
            ws.cell(row=row_num, column=5, value=participant.gender or "")
            ws.cell(row=row_num, column=6, value=participant.role.display_name if participant.role else "")
            ws.cell(row=row_num, column=7, value=participant.family.name if participant.family else "")
            ws.cell(row=row_num, column=8, value=participant.email or "")
            ws.cell(row=row_num, column=9, value=participant.phone or "")
            ws.cell(row=row_num, column=10, value=participant.final_price)
            ws.cell(row=row_num, column=11, value=total_paid)
            ws.cell(row=row_num, column=12, value=outstanding)
            ws.cell(row=row_num, column=13, value=participant.address or "")

        # Excel in Memory-Stream speichern
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        return excel_stream
