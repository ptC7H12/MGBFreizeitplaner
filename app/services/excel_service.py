"""Service für Excel-Exporte mit wiederverwendbaren Formatierungen"""
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet


class ExcelService:
    """
    Wiederverwendbare Excel-Export-Funktionen für konsistente Formatierung.

    Diese Klasse bietet zentrale Methoden für:
    - Header-Formatierung
    - Spaltenbreiten-Anpassung
    - Standard-Styles
    - Summenzeilen-Formatierung
    """

    # Standard-Farben
    HEADER_COLOR = "4472C4"  # Blau
    SUMMARY_COLOR = "D9E1F2"  # Hellblau
    GREEN_COLOR = "00B050"    # Grün für positive Werte
    RED_COLOR = "C00000"      # Rot für negative Werte
    WHITE_COLOR = "FFFFFF"    # Weiß

    @staticmethod
    def create_header_style() -> Dict[str, Any]:
        """
        Erstellt Standard Header-Formatierung.

        Returns:
            Dictionary mit 'fill', 'font' und 'alignment' Objekten
        """
        return {
            'fill': PatternFill(
                start_color=ExcelService.HEADER_COLOR,
                end_color=ExcelService.HEADER_COLOR,
                fill_type="solid"
            ),
            'font': Font(
                color=ExcelService.WHITE_COLOR,
                bold=True,
                size=11
            ),
            'alignment': Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
        }

    @staticmethod
    def create_summary_style() -> Dict[str, Any]:
        """
        Erstellt Standard Summenzeilen-Formatierung.

        Returns:
            Dictionary mit 'fill' und 'font' Objekten
        """
        return {
            'fill': PatternFill(
                start_color=ExcelService.SUMMARY_COLOR,
                end_color=ExcelService.SUMMARY_COLOR,
                fill_type="solid"
            ),
            'font': Font(bold=True)
        }

    @staticmethod
    def apply_header_row(
        ws: Worksheet,
        headers: List[str],
        column_widths: Optional[Dict[int, int]] = None,
        row: int = 1
    ) -> None:
        """
        Wendet Header-Formatierung auf eine Zeile an.

        Args:
            ws: Worksheet-Objekt
            headers: Liste der Header-Texte
            column_widths: Optional - Dictionary mit {Spalten-Index: Breite}
            row: Zeilennummer (default: 1)
        """
        style = ExcelService.create_header_style()

        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header_text)
            cell.fill = style['fill']
            cell.font = style['font']
            cell.alignment = style['alignment']

        # Spaltenbreiten anwenden falls angegeben
        if column_widths:
            for col_num, width in column_widths.items():
                column_letter = ws.cell(row=row, column=col_num).column_letter
                ws.column_dimensions[column_letter].width = width

    @staticmethod
    def apply_summary_row(
        ws: Worksheet,
        row: int,
        values: Dict[int, Any],
        label_column: int = 1,
        label_text: str = "GESAMT"
    ) -> None:
        """
        Wendet Summenzeilen-Formatierung an.

        Args:
            ws: Worksheet-Objekt
            row: Zeilennummer
            values: Dictionary mit {Spalten-Index: Wert}
            label_column: Spalte für Label (default: 1)
            label_text: Label-Text (default: "GESAMT")
        """
        style = ExcelService.create_summary_style()

        # Label setzen
        label_cell = ws.cell(row=row, column=label_column, value=label_text)
        label_cell.fill = style['fill']
        label_cell.font = style['font']

        # Werte setzen
        for col_num, value in values.items():
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.fill = style['fill']
            cell.font = style['font']

    @staticmethod
    def apply_color_by_value(
        ws: Worksheet,
        row: int,
        column: int,
        value: float,
        positive_color: Optional[str] = None,
        negative_color: Optional[str] = None
    ) -> None:
        """
        Färbt eine Zelle basierend auf positivem/negativem Wert.

        Args:
            ws: Worksheet-Objekt
            row: Zeilennummer
            column: Spaltennummer
            value: Wert zur Prüfung
            positive_color: Farbe für positive Werte (default: Grün)
            negative_color: Farbe für negative Werte (default: Rot)
        """
        if positive_color is None:
            positive_color = ExcelService.GREEN_COLOR
        if negative_color is None:
            negative_color = ExcelService.RED_COLOR

        cell = ws.cell(row=row, column=column)
        if value > 0:
            cell.font = Font(color=positive_color)
        elif value < 0:
            cell.font = Font(color=negative_color)

    @staticmethod
    def set_column_widths(
        ws: Worksheet,
        widths: List[int],
        start_column: int = 1
    ) -> None:
        """
        Setzt Spaltenbreiten für eine Liste von Spalten.

        Args:
            ws: Worksheet-Objekt
            widths: Liste von Breiten in derselben Reihenfolge wie Spalten
            start_column: Erste Spalte (default: 1)
        """
        for idx, width in enumerate(widths):
            col_num = start_column + idx
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

    @staticmethod
    def format_currency_cell(
        ws: Worksheet,
        row: int,
        column: int,
        value: float,
        color_by_sign: bool = False
    ) -> None:
        """
        Formatiert eine Zelle als Währung.

        Args:
            ws: Worksheet-Objekt
            row: Zeilennummer
            column: Spaltennummer
            value: Währungswert
            color_by_sign: Färbung nach Vorzeichen (default: False)
        """
        cell = ws.cell(row=row, column=column, value=value)
        cell.number_format = '#,##0.00'

        if color_by_sign:
            ExcelService.apply_color_by_value(ws, row, column, value)

    @staticmethod
    def create_workbook(sheet_title: str = "Sheet1") -> tuple[Workbook, Worksheet]:
        """
        Erstellt ein neues Workbook mit einer Worksheet.

        Args:
            sheet_title: Titel des Worksheets

        Returns:
            Tuple aus (Workbook, Worksheet)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        return wb, ws
