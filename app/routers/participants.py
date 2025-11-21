"""Participants (Teilnehmer) Router"""
import logging
import json
import csv
from io import BytesIO, StringIO
from fastapi import APIRouter, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError, DataError
from datetime import date, datetime
from typing import Optional, List, Dict, Any
from pydantic import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from app.database import get_db, transaction
from app.models import Participant, Role, Event, Family, Ruleset, Setting, Task
from app.services.price_calculator import PriceCalculator
from app.services.qrcode_service import QRCodeService
from app.services.excel_service import ExcelService
from app.dependencies import get_current_event_id
from app.utils.error_handler import handle_db_exception
from app.utils.flash import flash
from app.utils.datetime_utils import utcnow
from app.schemas import ParticipantCreateSchema, ParticipantUpdateSchema
from app.templates_config import templates

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/participants", tags=["participants"])


def _calculate_price_for_participant(
    db: Session,
    event_id: int,
    role_id: Optional[int],  # Rolle ist optional
    birth_date: date,
    family_id: Optional[int]
) -> float:
    """
    Wrapper-Funktion für Preisberechnung (ruft PriceCalculator.calculate_price_from_db auf)

    Diese Funktion ist für Abwärtskompatibilität erhalten, delegiert aber an den
    PriceCalculator-Service, um Code-Duplikation zu vermeiden.

    Args:
        db: Datenbank-Session
        event_id: ID der Veranstaltung
        role_id: Optional ID der Rolle (kann None sein)
        birth_date: Geburtsdatum des Teilnehmers
        family_id: Optional ID der Familie

    Returns:
        Berechneter Preis in Euro
    """
    return PriceCalculator.calculate_price_from_db(
        db=db,
        event_id=event_id,
        role_id=role_id,
        birth_date=birth_date,
        family_id=family_id
    )


def _check_and_update_role_count_task(
    db: Session,
    event_id: int,
    role_id: Optional[int]
) -> None:
    """
    Prüft ob die Rollenanzahl überschritten wurde und erstellt/aktualisiert/schließt die entsprechende Task.

    Wird nach dem Erstellen, Bearbeiten oder Löschen eines Teilnehmers aufgerufen.

    Args:
        db: Datenbank-Session
        event_id: ID der Veranstaltung
        role_id: ID der Rolle (kann None sein)
    """
    if not role_id:
        return  # Keine Rolle zugewiesen, nichts zu prüfen

    # Hole die Rolle mit allen Informationen
    role = db.query(Role).filter(Role.id == role_id, Role.event_id == event_id).first()
    if not role:
        return

    # Hole das aktive Ruleset für dieses Event
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True
    ).first()

    if not ruleset or not ruleset.role_discounts:
        return  # Kein Ruleset oder keine Rollenrabatte konfiguriert

    # Prüfe ob diese Rolle ein max_count hat
    role_config = ruleset.role_discounts.get(role.name)
    if not role_config:
        return  # Rolle nicht im Ruleset gefunden

    max_count = role_config.get("max_count")
    if max_count is None:
        return  # Keine Begrenzung für diese Rolle

    # Zähle aktive Teilnehmer mit dieser Rolle
    current_count = db.query(Participant).filter(
        Participant.event_id == event_id,
        Participant.role_id == role.id,
        Participant.is_active == True
    ).count()

    # Suche nach bestehender Task für diese Rolle
    existing_task = db.query(Task).filter(
        Task.event_id == event_id,
        Task.task_type == "role_count_exceeded",
        Task.reference_id == role.id
    ).first()

    # Überschreitung?
    if current_count > max_count:
        excess_count = current_count - max_count
        description = f"Aktuell: {current_count} | Maximum: {max_count} | Überschreitung: {excess_count}"

        if existing_task:
            # Task existiert bereits - aktualisiere sie
            existing_task.description = description
            existing_task.is_completed = False  # Wieder öffnen, falls sie abgeschlossen war
            existing_task.updated_at = utcnow()
            logger.info(f"Role count task updated and reopened for {role.display_name}: {description}")
        else:
            # Task existiert noch nicht - erstelle sie
            new_task = Task(
                event_id=event_id,
                task_type="role_count_exceeded",
                reference_id=role.id,
                title=f"Zu viele {role.display_name} zugewiesen",
                description=description,
                is_completed=False
            )
            db.add(new_task)
            logger.info(f"Role count task created for {role.display_name}: {description}")
    else:
        # Keine Überschreitung mehr
        if existing_task and not existing_task.is_completed:
            # Task existiert und ist offen - schließe sie automatisch
            existing_task.is_completed = True
            existing_task.updated_at = utcnow()
            logger.info(f"Role count task auto-completed for {role.display_name}: within limit ({current_count}/{max_count})")

    # Änderungen werden durch den äußeren db.commit() gespeichert


@router.get("/", response_class=HTMLResponse)
async def list_participants(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    search: Optional[str] = "",
    role_id: Optional[str] = "",
    payment_status: Optional[str] = "",
    family_search: Optional[str] = "",
    family_role_id: Optional[str] = "",
    family_payment_status: Optional[str] = ""
):
    """Liste aller Teilnehmer mit Such- und Filterfunktionen"""
    query = db.query(Participant).filter(Participant.event_id == event_id)

    # Suchfilter
    if search and search.strip():
        search_filter = f"%{search}%"
        query = query.filter(
            (Participant.first_name.ilike(search_filter)) |
            (Participant.last_name.ilike(search_filter)) |
            (Participant.email.ilike(search_filter))
        )

    # Rollenfilter (nur wenn nicht leer)
    if role_id and role_id.strip():
        try:
            query = query.filter(Participant.role_id == int(role_id))
        except ValueError:
            pass  # Ungültige ID ignorieren

    # Eager Loading für related objects um N+1 Queries zu vermeiden
    participants = query.options(
        joinedload(Participant.role),
        joinedload(Participant.family),
        joinedload(Participant.event),
        joinedload(Participant.payments)
    ).order_by(Participant.last_name).all()

    # Berechne Zahlungsinformationen für jeden Teilnehmer
    participant_data = []
    for participant in participants:
        # Direkte Zahlungen des Teilnehmers
        direct_payments = float(sum((payment.amount for payment in participant.payments), 0))

        # Anteilige Familienzahlungen berechnen
        family_payment_share = 0.0
        if participant.family_id:
            # Lade Familie mit Zahlungen und Teilnehmern
            family = db.query(Family).filter(Family.id == participant.family_id).options(
                joinedload(Family.payments),
                joinedload(Family.participants).joinedload(Participant.payments)
            ).first()

            if family:
                # Gesamtzahlungen der Familie
                family_total_payments = float(sum((payment.amount for payment in family.payments), 0))

                # Berechne für jedes Familienmitglied die direkten Zahlungen und offenen Beträge
                family_members_outstanding = []
                total_outstanding = 0.0

                for member in family.participants:
                    member_direct_payments = float(sum((payment.amount for payment in member.payments), 0))
                    member_outstanding = max(0.0, float(member.final_price) - member_direct_payments)
                    family_members_outstanding.append({
                        'participant_id': member.id,
                        'outstanding': member_outstanding
                    })
                    total_outstanding += member_outstanding

                # Verteile die Familienzahlung proportional auf die offenen Beträge
                if total_outstanding > 0:
                    # Finde den aktuellen Teilnehmer in der Liste
                    for member_data in family_members_outstanding:
                        if member_data['participant_id'] == participant.id:
                            # Anteilige Verteilung basierend auf offenen Beträgen
                            family_payment_share = (member_data['outstanding'] / total_outstanding) * family_total_payments
                            break
                elif family_total_payments > 0:
                    # Wenn alle Beträge bezahlt sind, aber noch Familienzahlungen da sind,
                    # verteile proportional nach Sollpreis
                    family_total_price = float(sum((p.final_price for p in family.participants), 0))
                    if family_total_price > 0:
                        family_payment_share = (float(participant.final_price) / family_total_price) * family_total_payments

        # Gesamtzahlung = direkte Zahlungen + anteilige Familienzahlungen
        total_paid = direct_payments + family_payment_share
        outstanding = float(participant.final_price) - total_paid

        # Zahlungsstatus-Filter anwenden
        if payment_status == "paid" and outstanding > 0.01:
            continue  # Überspringe nicht vollständig bezahlte
        elif payment_status == "outstanding" and outstanding <= 0.01:
            continue  # Überspringe vollständig bezahlte

        participant_data.append({
            "participant": participant,
            "total_paid": total_paid,
            "outstanding": outstanding
        })

    # Für Filter-Dropdown (auch nach event_id gefiltert)
    roles = db.query(Role).filter(Role.is_active == True, Role.event_id == event_id).all()
    families = db.query(Family).filter(Family.event_id == event_id).order_by(Family.name).all()

    # Für Familien-Tab: Familiendaten mit Statistiken berechnen
    families_with_participants = db.query(Family)\
        .filter(Family.event_id == event_id)\
        .options(
            joinedload(Family.participants),
            joinedload(Family.payments)
        )\
        .order_by(Family.name)\
        .all()

    family_data = []
    for family in families_with_participants:
        # Suchfilter für Familien
        if family_search and family_search.strip():
            search_filter = family_search.strip().lower()
            family_name = (family.name or "").lower()
            contact_person = (family.contact_person or "").lower()
            if search_filter not in family_name and search_filter not in contact_person:
                continue  # Überspringe Familien, die nicht zur Suche passen

        # Rollenfilter für Familien - prüfe ob ein Teilnehmer der Familie diese Rolle hat
        if family_role_id and family_role_id.strip():
            try:
                role_id_int = int(family_role_id)
                has_role = any(p.role_id == role_id_int for p in family.participants)
                if not has_role:
                    continue  # Überspringe Familien ohne Teilnehmer mit dieser Rolle
            except ValueError:
                pass

        # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
        total_price = float(sum((p.final_price for p in family.participants), 0))

        # Zahlungen: Sowohl direkte Familienzahlungen als auch Zahlungen an einzelne Mitglieder
        family_payments = float(sum((payment.amount for payment in family.payments), 0))
        member_payments = float(sum(
            (payment.amount
            for participant in family.participants
            for payment in participant.payments), 0
        ))
        total_paid = family_payments + member_payments
        outstanding = total_price - total_paid

        # Zahlungsstatus-Filter für Familien
        if family_payment_status == "paid" and outstanding > 0.01:
            continue  # Überspringe nicht vollständig bezahlte Familien
        elif family_payment_status == "outstanding" and outstanding <= 0.01:
            continue  # Überspringe vollständig bezahlte Familien

        family_data.append({
            "family": family,
            "participant_count": len(family.participants),
            "total_price": total_price,
            "total_paid": total_paid,
            "outstanding": outstanding
        })

    return templates.TemplateResponse(
        "participants/list.html",
        {
            "request": request,
            "title": "Teilnehmer & Familien",
            "participant_data": participant_data,
            "roles": roles,
            "families": families,
            "family_data": family_data,
            "search": search,
            "selected_role_id": role_id,
            "payment_status": payment_status,
            "family_search": family_search,
            "family_role_id": family_role_id,
            "family_payment_status": family_payment_status
        }
    )


@router.get("/create", response_class=HTMLResponse)
async def create_participant_form(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Formular zum Erstellen eines neuen Teilnehmers"""
    roles = db.query(Role).filter(Role.is_active == True, Role.event_id == event_id).all()
    event = db.query(Event).filter(Event.id == event_id).first()
    families = db.query(Family).filter(Family.event_id == event_id).order_by(Family.name).all()

    return templates.TemplateResponse(
        "participants/create.html",
        {
            "request": request,
            "title": "Neuer Teilnehmer",
            "roles": roles,
            "event": event,
            "families": families
        }
    )


@router.post("/create", response_class=HTMLResponse)
async def create_participant(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    first_name: str = Form(...),
    last_name: str = Form(...),
    birth_date: str = Form(...),
    gender: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    bildung_teilhabe_id: Optional[str] = Form(None),
    allergies: Optional[str] = Form(None),
    medical_notes: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    discount_percent: str = Form("0.0"),  # Als String empfangen
    discount_reason: Optional[str] = Form(None),
    manual_price_override: str = Form(""),  # Als String empfangen
    role_id: Optional[str] = Form(None),  # Als String empfangen optional
    family_id: Optional[str] = Form(None),  # Als String empfangen
    create_as_family: Optional[str] = Form(None)
):
    """Erstellt einen neuen Teilnehmer"""
    try:
        # Konvertiere leere Strings zu None oder passenden Werten
        email_val = None if not email or email.strip() == "" else email
        manual_price_override_val = None if not manual_price_override or manual_price_override.strip() == "" else float(manual_price_override)
        discount_percent_val = 0.0 if not discount_percent or discount_percent.strip() == "" else float(discount_percent)
        role_id_val = None if not role_id or role_id.strip() == "" else int(role_id)
        family_id_val = None if not family_id or family_id.strip() == "" else int(family_id)

        # Pydantic-Validierung
        participant_data = ParticipantCreateSchema(
            first_name=first_name,
            last_name=last_name,
            birth_date=birth_date,
            gender=gender,
            email=email_val,
            phone=phone,
            address=address,
            bildung_teilhabe_id=bildung_teilhabe_id,
            allergies=allergies,
            medical_notes=medical_notes,
            notes=notes,
            discount_percent=discount_percent_val,
            discount_reason=discount_reason,
            manual_price_override=manual_price_override_val,
            event_id=event_id,  # From session dependency
            role_id=role_id_val,
            family_id=family_id_val
        )

        # Wenn "Als Familie erstellen" aktiviert ist, neue Familie erstellen
        if create_as_family == "true":
            # Familienname aus Nachname + Vorname bilden (wegen möglicher Duplikate)
            family_name = f"{last_name} {first_name}"

            # Prüfen ob Familie mit diesem Namen schon existiert
            existing_family = db.query(Family).filter(
                Family.event_id == event_id,
                Family.name == family_name
            ).first()

            if existing_family:
                # Familie existiert bereits, diese verwenden
                family_id_val = existing_family.id
            else:
                # Neue Familie erstellen
                new_family = Family(
                    name=family_name,
                    event_id=event_id,
                    contact_person=f"{first_name} {last_name}",
                    email=email if email else None,
                    phone=phone if phone else None,
                    address=address if address else None
                )
                db.add(new_family)
                db.flush()  # Familie speichern, um ID zu erhalten
                family_id_val = new_family.id

        # Automatische Preisberechnung (birth_date ist bereits ein date-Objekt)
        # Verwende family_id_val (kann durch "Als Familie erstellen" gesetzt sein)
        calculated_price = _calculate_price_for_participant(
            db=db,
            event_id=event_id,
            role_id=participant_data.role_id,
            birth_date=participant_data.birth_date,
            family_id=family_id_val
        )

        # Neuen Teilnehmer erstellen (birth_date ist bereits ein date-Objekt)
        participant = Participant(
            first_name=participant_data.first_name,
            last_name=participant_data.last_name,
            birth_date=participant_data.birth_date,
            gender=participant_data.gender,
            email=participant_data.email,
            phone=participant_data.phone,
            address=participant_data.address,
            bildung_teilhabe_id=participant_data.bildung_teilhabe_id,
            allergies=participant_data.allergies,
            medical_notes=participant_data.medical_notes,
            notes=participant_data.notes,
            discount_percent=participant_data.discount_percent,
            discount_reason=participant_data.discount_reason,
            manual_price_override=participant_data.manual_price_override,
            event_id=event_id,  # Aus Session, nicht aus Formular!
            role_id=participant_data.role_id,
            family_id=family_id_val,  # Verwende konvertierte Variable
            calculated_price=calculated_price
        )

        # Transaction Context Manager: Auto-commit bei Erfolg, auto-rollback bei Exception
        with transaction(db):
            db.add(participant)
            db.flush()  # Generiert ID ohne zu committen
        # Auto-commit erfolgt hier

        # Prüfe Rollenüberschreitung und aktualisiere Task
        _check_and_update_role_count_task(db, event_id, participant.role_id)
        db.commit()  # Speichere Task-Änderungen

        flash(request, f"Teilnehmer {participant.full_name} wurde erfolgreich erstellt", "success")
        return RedirectResponse(url=f"/participants/{participant.id}", status_code=303)

    except ValidationError as e:
        # Pydantic-Validierungsfehler
        logger.warning(f"Validation error creating participant: {e}", exc_info=True)
        # Ersten Fehler extrahieren für benutzerfreundliche Nachricht
        first_error = e.errors()[0]
        field_name = first_error['loc'][0] if first_error['loc'] else 'Unbekannt'
        error_msg = first_error['msg']
        flash(request, f"Validierungsfehler ({field_name}): {error_msg}", "error")
        return RedirectResponse(url="/participants/create?error=validation", status_code=303)

    except IntegrityError as e:
        db.rollback()
        logger.exception(f"Database integrity error creating participant: {e}")
        flash(request, "Teilnehmer konnte nicht erstellt werden (Datenbankfehler)", "error")
        return RedirectResponse(url="/participants/create?error=db_integrity", status_code=303)

    except DataError as e:
        db.rollback()
        logger.exception(f"Invalid data creating participant: {e}")
        flash(request, "Ungültige Daten eingegeben", "error")
        return RedirectResponse(url="/participants/create?error=invalid_data", status_code=303)

    except Exception as e:
        return handle_db_exception(e, "/participants/create", "Creating participant", db, request)


@router.post("/calculate-price", response_class=HTMLResponse)
async def calculate_price_preview(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    birth_date: str = Form(...),
    role_id: int = Form(...),
    family_id: Optional[int] = Form(None)
):
    """
    HTMX-Endpunkt für Live-Preis-Vorschau
    Berechnet den Preis basierend auf Geburtsdatum, Event, Rolle und Familie
    """
    try:
        # Datum parsen
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()

        # Preis berechnen
        calculated_price = _calculate_price_for_participant(
            db=db,
            event_id=event_id,
            role_id=role_id,
            birth_date=birth_date_obj,
            family_id=family_id
        )

        # Event und Rolle laden für Detailinfo
        event = db.query(Event).filter(Event.id == event_id).first()
        role = db.query(Role).filter(Role.id == role_id).first()

        # Alter berechnen
        age = None
        if event:
            age = event.start_date.year - birth_date_obj.year
            if (event.start_date.month, event.start_date.day) < (birth_date_obj.month, birth_date_obj.day):
                age -= 1

        # HTML-Fragment zurückgeben
        return templates.TemplateResponse(
            "participants/_price_preview.html",
            {
                "request": request,
                "calculated_price": calculated_price,
                "age": age,
                "event_name": event.name if event else None,
                "role_name": role.name if role else None
            }
        )

    except ValueError as e:
        logger.warning(f"Invalid date input for price calculation: {e}", exc_info=True)
        return HTMLResponse(content=f'<div class="text-sm text-red-500">Ungültiges Datum</div>')

    except Exception as e:
        logger.exception(f"Error calculating price preview: {e}")
        return HTMLResponse(content=f'<div class="text-sm text-gray-500">Preis wird berechnet...</div>')


@router.post("/suggest-role", response_class=HTMLResponse)
async def suggest_role(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    birth_date: str = Form(...)
):
    """
    HTMX-Endpunkt der die passende Rolle basierend auf dem Alter vorschlägt
    """
    try:
        # Datum parsen
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()

        # Event laden
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            return HTMLResponse(content="")

        # Alter berechnen
        age = event.start_date.year - birth_date_obj.year
        if (event.start_date.month, event.start_date.day) < (birth_date_obj.month, birth_date_obj.day):
            age -= 1

        # Aktives Regelwerk finden
        ruleset = db.query(Ruleset).filter(
            Ruleset.is_active == True,
            Ruleset.valid_from <= event.start_date,
            Ruleset.valid_until >= event.start_date
        ).first()

        if not ruleset or not ruleset.age_groups:
            return HTMLResponse(content="")

        # Passende Altersgruppe finden
        suggested_role_name = None
        for age_group in ruleset.age_groups:
            min_age = age_group.get("min_age", 0)
            max_age = age_group.get("max_age", 999)
            if min_age <= age <= max_age:
                suggested_role_name = age_group.get("role", "").lower()
                break

        if not suggested_role_name:
            return HTMLResponse(content="")

        # Rolle anhand des Namens finden
        roles = db.query(Role).filter(
            Role.event_id == event_id,
            Role.is_active == True
        ).all()

        suggested_role_id = None
        for role in roles:
            if role.name.lower() == suggested_role_name:
                suggested_role_id = role.id
                break

        if suggested_role_id:
            # JavaScript zurückgeben, das die Rolle auswählt
            return HTMLResponse(content=f"""
                <script>
                    document.getElementById('role_id').value = '{suggested_role_id}';
                    // Trigger price calculation
                    htmx.trigger('#role_id', 'change');
                </script>
            """)
        else:
            return HTMLResponse(content="")

    except ValueError as e:
        logger.warning(f"Invalid date input for role suggestion: {e}", exc_info=True)
        return HTMLResponse(content="")

    except Exception as e:
        logger.exception(f"Error suggesting role: {e}")
        return HTMLResponse(content="")


# ===== IMPORT/EXPORT ROUTES (müssen VOR /{participant_id} definiert werden) =====
# Diese Routes haben spezifische Pfade und müssen vor parametrisierten Routes kommen



@router.get("/import", response_class=HTMLResponse)
async def import_participants_form(
    request: Request,
    db: Session = Depends(get_db)
):
    """Zeigt das Excel-Import Formular"""
    return templates.TemplateResponse(
        "participants/import.html",
        {
            "request": request,
            "title": "Excel Import"
        }
    )


@router.get("/import/template")
async def download_import_template(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    format: str = "xlsx"
):
    """Generiert eine Excel- oder CSV-Vorlage zum Herunterladen"""

    headers = [
        "Vorname*", "Nachname*", "Geburtsdatum* (TT.MM.JJJJ)",
        "Geschlecht", "E-Mail", "Telefon", "Adresse", "Familien-Nr"
    ]

    example_data = [
        ["Max", "Mustermann", "15.03.2010", "männlich", "max@example.com", "0123456789", "Musterstraße 1, 12345 Stadt", "1"],
        ["Maria", "Mustermann", "20.07.2012", "weiblich", "maria@example.com", "", "Musterstraße 1, 12345 Stadt", "1"],
        ["Anna", "Schmidt", "05.09.2011", "weiblich", "anna@example.com", "0987654321", "", ""]
    ]

    # CSV-Format
    if format == "csv":
        csv_buffer = StringIO()
        csv_writer = csv.writer(csv_buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        # Header schreiben
        csv_writer.writerow(headers)

        # Beispieldaten schreiben
        for row in example_data:
            csv_writer.writerow(row)

        csv_content = csv_buffer.getvalue()

        return Response(
            content=csv_content.encode('utf-8-sig'),  # BOM für Excel-Kompatibilität
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=Teilnehmer_Import_Vorlage.csv"
            }
        )

    # Excel-Format (Standard) mit ExcelService
    wb, ws = ExcelService.create_workbook("Teilnehmer")

    # Spaltenbreiten
    column_widths = {1: 15, 2: 15, 3: 25, 4: 12, 5: 25, 6: 15, 7: 30, 8: 12}

    # Header-Formatierung mit ExcelService
    ExcelService.apply_header_row(ws, headers, column_widths)

    # Beispieldaten hinzufügen (bereits definiert oben)

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Hinweise in separatem Sheet
    ws_info = wb.create_sheet("Hinweise")
    ws_info.column_dimensions['A'].width = 80

    info_text = [
        "Hinweise zum Excel-Import:",
        "",
        "PFLICHTFELDER (mit * markiert):",
        "- Vorname: Vorname des Teilnehmers",
        "- Nachname: Nachname des Teilnehmers",
        "- Geburtsdatum: Format TT.MM.JJJJ (z.B. 15.03.2010)",
        "",
        "OPTIONALE FELDER:",
        "- Geschlecht: männlich/weiblich/divers oder leer lassen",
        "- E-Mail: E-Mail-Adresse des Teilnehmers oder Erziehungsberechtigten",
        "- Telefon: Telefonnummer",
        "- Adresse: Vollständige Adresse",
        "- Familien-Nr: Gleiche Nummer für Familienmitglieder (z.B. 1, 2, 3...)",
        "",
        "FAMILIEN-GRUPPIERUNG:",
        "Wenn mehrere Teilnehmer zur gleichen Familie gehören, geben Sie",
        "die gleiche Familien-Nummer ein (z.B. alle '1', dann nächste Familie '2', etc.)",
        "Der Familienname wird automatisch aus 'Nachname Vorname' des ersten",
        "Mitglieds erstellt.",
        "",
        "BEISPIEL:",
        "Max Mustermann (Familie 1)",
        "Maria Mustermann (Familie 1)  <- Gleiche Familie",
        "Anna Schmidt (Familie 2 oder leer)  <- Andere/Keine Familie"
    ]

    for row_num, text in enumerate(info_text, 1):
        cell = ws_info.cell(row=row_num, column=1, value=text)
        if "PFLICHTFELDER" in text or "OPTIONALE" in text or "FAMILIEN" in text:
            cell.font = Font(bold=True, size=12)

    # Excel in BytesIO speichern
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return Response(
        content=excel_buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Teilnehmer_Import_Vorlage.xlsx"
        }
    )


def _parse_csv_data(csv_content: str) -> List[List[str]]:
    """
    Hilfsfunktion zum Parsen von CSV-Inhalten.

    Args:
        csv_content: CSV-Inhalt als String

    Returns:
        Liste von Zeilen, jede Zeile ist eine Liste von Werten
    """
    csv_reader = csv.reader(StringIO(csv_content), delimiter=';')
    rows = list(csv_reader)

    # Fallback: Versuche Komma als Delimiter wenn Semikolon keine Spalten ergibt
    if rows and len(rows[0]) <= 1:
        csv_reader = csv.reader(StringIO(csv_content), delimiter=',')
        rows = list(csv_reader)

    return rows


def _process_import_row(
    row: List[Any],
    row_num: int,
    participants_data: List[Dict[str, Any]],
    errors: List[str],
    families_dict: Dict[str, Family]
) -> None:
    """
    Hilfsfunktion zum Verarbeiten einer Import-Zeile (Excel oder CSV).

    Args:
        row: Zeile mit Teilnehmerdaten
        row_num: Zeilennummer für Fehlermeldungen
        participants_data: Liste zum Sammeln der verarbeiteten Teilnehmerdaten
        errors: Liste zum Sammeln von Fehlermeldu ngen
        families_dict: Dictionary mit Familien-Nummern als Keys
    """
    # Leere Zeilen überspringen
    if not any(row):
        return

    # Werte aus der Zeile extrahieren
    first_name = str(row[0]).strip() if row[0] else ""
    last_name = str(row[1]).strip() if row[1] else ""

    # Geburtsdatum: Kann als date/datetime-Objekt oder String kommen
    birth_date_raw = row[2] if len(row) > 2 and row[2] else None

    gender = str(row[3]).strip() if row[3] and len(row) > 3 else ""
    email = str(row[4]).strip() if row[4] and len(row) > 4 else ""
    phone = str(row[5]).strip() if row[5] and len(row) > 5 else ""
    address = str(row[6]).strip() if row[6] and len(row) > 6 else ""

    # Familien-Nummer: Konvertiere Zahlen zu Int dann String (verhindert 1.0 vs "1" Problem)
    family_number = ""
    if len(row) > 7 and row[7]:
        if isinstance(row[7], (int, float)):
            # Excel gibt manchmal Zahlen als Float zurück (1.0) - konvertiere zu Int für Konsistenz
            family_number = str(int(row[7]))
        else:
            family_number = str(row[7]).strip()

    # Validierung
    row_errors = []
    has_error = False

    if not first_name:
        row_errors.append("Vorname fehlt")
        has_error = True

    if not last_name:
        row_errors.append("Nachname fehlt")
        has_error = True

    # Geburtsdatum parsen - unterstützt mehrere Formate
    birth_date = None
    birth_date_str = ""  # Für Fehlerausgabe
    
    if birth_date_raw:
        try:
            # Fall 1: Excel gibt bereits ein date/datetime-Objekt zurück
            if isinstance(birth_date_raw, date):
                birth_date = birth_date_raw
                birth_date_str = birth_date.strftime("%d.%m.%Y")
            elif isinstance(birth_date_raw, datetime):
                birth_date = birth_date_raw.date()
                birth_date_str = birth_date.strftime("%d.%m.%Y")
            else:
                # Fall 2: String parsen - versuche verschiedene Formate
                birth_date_str = str(birth_date_raw).strip()
                
                # Unterstützte Formate
                date_formats = [
                    "%Y-%m-%d %H:%M:%S",  # Excel: 1991-07-01 00:00:00
                    "%Y-%m-%d",           # ISO: 1991-07-01
                    "%d.%m.%Y",           # Deutsch: 01.07.1991
                    "%d/%m/%Y",           # Alternative: 01/07/1991
                    "%d-%m-%Y",           # Alternative: 01-07-1991
                ]
                
                for fmt in date_formats:
                    try:
                        birth_date = datetime.strptime(birth_date_str, fmt).date()
                        break
                    except ValueError:
                        continue

                if not birth_date:
                    row_errors.append(f"Ungültiges Datumsformat: {birth_date_str}")
                    has_error = True
                    
        except Exception as e:
            row_errors.append(f"Fehler beim Datum: {str(e)}")
            has_error = True
    else:
        row_errors.append("Geburtsdatum fehlt")
        has_error = True

    if row_errors:
        errors.append({
            "row": row_num,
            "message": ", ".join(row_errors)
        })

    # Teilnehmer zur Liste hinzufügen
    participant_data = {
        "first_name": first_name,
        "last_name": last_name,
        "birth_date": birth_date.strftime("%d.%m.%Y") if birth_date else birth_date_str,
        "birth_date_obj": birth_date,
        "gender": gender if gender else None,
        "email": email if email else None,
        "phone": phone if phone else None,
        "address": address if address else None,
        "family_number": family_number if family_number else None,
        "has_error": has_error,
        "row": row_num
    }

    participants_data.append(participant_data)

    # Familien gruppieren
    if family_number and not has_error:
        if family_number not in families_dict:
            families_dict[family_number] = []
        families_dict[family_number].append(participant_data)


@router.post("/import", response_class=HTMLResponse)
async def upload_import_file(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Verarbeitet die hochgeladene Excel- oder CSV-Datei und zeigt eine Vorschau"""
    try:
        # Datei validieren
        is_csv = file.filename.endswith('.csv')
        is_excel = file.filename.endswith(('.xlsx', '.xls'))

        if not (is_csv or is_excel):
            flash(request, "Bitte laden Sie eine Excel-Datei (.xlsx, .xls) oder CSV-Datei (.csv) hoch", "error")
            return RedirectResponse(url="/participants/import", status_code=303)

        # Datei lesen
        contents = await file.read()

        # Daten parsen
        participants_data = []
        errors = []
        families_dict = {}  # family_number -> [participants]

        # CSV-Datei
        if is_csv:
            csv_content = contents.decode('utf-8-sig')  # UTF-8 mit BOM Support
            rows = _parse_csv_data(csv_content)

            if not rows or len(rows) < 2:
                flash(request, "CSV-Datei ist leer oder enthält keine Daten", "error")
                return RedirectResponse(url="/participants/import", status_code=303)

            # Header ist Zeile 0, Daten ab Zeile 1
            for row_num, row in enumerate(rows[1:], start=2):
                _process_import_row(row, row_num, participants_data, errors, families_dict)

        # Excel-Datei
        else:
            wb = load_workbook(BytesIO(contents), data_only=True)
            ws = wb.active

            # Header prüfen (erste Zeile)
            header_row = [cell.value for cell in ws[1]]

            # Ab Zeile 2 lesen (Zeile 1 ist Header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                _process_import_row(row, row_num, participants_data, errors, families_dict)

        if not participants_data:
            flash(request, "Keine Teilnehmer in der Datei gefunden", "error")
            return RedirectResponse(url="/participants/import", status_code=303)

        # Import-Daten für Vorschau
        import_data = {
            "participants": participants_data,
            "families": families_dict,
            "errors": errors
        }

        # Als JSON für das Formular
        import_data_json = json.dumps(import_data, default=str)

        return templates.TemplateResponse(
            "participants/import_preview.html",
            {
                "request": request,
                "title": "Import-Vorschau",
                "import_data": import_data,
                "import_data_json": import_data_json
            }
        )

    except Exception as e:
        logger.exception(f"Error processing Excel file: {e}")
        flash(request, f"Fehler beim Verarbeiten der Datei: {str(e)}", "error")
        return RedirectResponse(url="/participants/import", status_code=303)


@router.post("/import/confirm", response_class=HTMLResponse)
async def confirm_import(
    request: Request,
    import_data: str = Form(...),
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Führt den eigentlichen Import durch"""
    try:
        # Import-Daten deserialisieren
        data = json.loads(import_data)

        imported_count = 0
        skipped_count = 0
        family_map = {}  # family_number -> family_id

        # Event laden
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            flash(request, "Event nicht gefunden", "error")
            return RedirectResponse(url="/participants/import", status_code=303)

        # Aktives Ruleset laden für Preisberechnung
        ruleset = db.query(Ruleset).filter(
            Ruleset.is_active == True,
            Ruleset.valid_from <= event.start_date,
            Ruleset.valid_until >= event.start_date
        ).first()

        # Zuerst: Familien erstellen
        for family_number, members in data.get("families", {}).items():
            if not members:
                continue

            # Ersten Teilnehmer der Familie nehmen für Familiennamen (sortiert nach Zeilennummer)
            # Damit ist sichergestellt, dass der erste in der Excel/CSV auch als erster verwendet wird
            sorted_members = sorted(members, key=lambda m: m.get('row', 0))
            first_member = sorted_members[0]
            family_name = f"Familie {first_member['last_name']} {first_member['first_name']}"

            # Familie erstellen
            new_family = Family(
                name=family_name,
                event_id=event_id,
                contact_person=f"{first_member['first_name']} {first_member['last_name']}",
                email=first_member.get('email'),
                phone=first_member.get('phone'),
                address=first_member.get('address')
            )

            db.add(new_family)
            db.flush()  # Familie speichern um ID zu bekommen

            family_map[family_number] = new_family.id

        # Dann: Teilnehmer importieren
        logger.info(f"Starting participant import: {len(data.get('participants', []))} participants to process")
        
        for participant_data in data.get("participants", []):
            # Debug-Logging
            logger.debug(f"Processing: {participant_data.get('first_name')} {participant_data.get('last_name')} - has_error: {participant_data.get('has_error')}")
            
            # Fehlerhafte überspringen
            if participant_data.get("has_error"):
                logger.warning(f"Skipping {participant_data.get('first_name')} {participant_data.get('last_name')} due to validation errors")
                skipped_count += 1
                continue
        
            try:
                # Geburtsdatum parsen
                birth_date_str = participant_data["birth_date"]
                birth_date = None
                
                # Versuche erst mit birth_date_obj (wenn bereits geparst)
                if participant_data.get("birth_date_obj"):
                    birth_date_obj = participant_data["birth_date_obj"]
                    
                    # Kann String oder date-Objekt sein (durch JSON-Serialisierung)
                    if isinstance(birth_date_obj, date):
                        birth_date = birth_date_obj
                        logger.debug(f"Using pre-parsed birth_date_obj (date): {birth_date}")
                    elif isinstance(birth_date_obj, datetime):
                        birth_date = birth_date_obj.date()
                        logger.debug(f"Using pre-parsed birth_date_obj (datetime): {birth_date}")
                    elif isinstance(birth_date_obj, str):
                        # JSON hat es zu String gemacht - zurück parsen
                        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]:
                            try:
                                parsed = datetime.strptime(birth_date_obj, fmt)
                                birth_date = parsed.date() if isinstance(parsed, datetime) else parsed
                                logger.debug(f"Parsed birth_date_obj string with format {fmt}: {birth_date}")
                                break
                            except ValueError:
                                continue
                
                # Falls birth_date_obj nicht funktioniert hat, versuche birth_date String
                if not birth_date and birth_date_str:
                    for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"]:
                        try:
                            parsed = datetime.strptime(birth_date_str, fmt)
                            birth_date = parsed.date() if isinstance(parsed, datetime) else parsed
                            logger.debug(f"Parsed birth_date string with format {fmt}: {birth_date}")
                            break
                        except ValueError:
                            continue
                
                if not birth_date:
                    logger.error(f"Could not parse birth_date for {participant_data.get('first_name')} {participant_data.get('last_name')}: {birth_date_str}")
                    skipped_count += 1
                    continue
                
                # Sicherstellen dass birth_date ein date-Objekt ist
                if isinstance(birth_date, datetime):
                    birth_date = birth_date.date()
                
                logger.debug(f"Final birth_date (type: {type(birth_date).__name__}): {birth_date}")
        
                # Rolle basierend auf Alter ermitteln
                age = event.start_date.year - birth_date.year
                if (event.start_date.month, event.start_date.day) < (birth_date.month, birth_date.day):
                    age -= 1
        
                logger.debug(f"Calculated age: {age} for {participant_data.get('first_name')} {participant_data.get('last_name')}")
        
                # KEINE automatische Rollenzuweisung beim Import
                # Teilnehmer bekommen nur eine Rolle, wenn sie explizit eine benötigen
                role_id = None
                logger.debug(f"Import: No automatic role assignment for {participant_data.get('first_name')} {participant_data.get('last_name')}")

                # Familie zuordnen
                family_id = None
                family_number = participant_data.get("family_number")
                if family_number and family_number in family_map:
                    family_id = family_map[family_number]
                    logger.debug(f"Assigned to family_id: {family_id}")

                # Preis berechnen: Nur Basispreis basierend auf Alter (aus age_groups)
                # Familienrabatte werden später vom PriceCalculator berechnet
                final_price = 0.0
                if ruleset and ruleset.age_groups:
                    # Finde passende Altersgruppe
                    for age_group in ruleset.age_groups:
                        min_age = age_group.get("min_age", 0)
                        max_age = age_group.get("max_age", 999)
                        if min_age <= age <= max_age:
                            # YAML verwendet "price", nicht "base_price"
                            base_price = age_group.get("price", 0.0)
                            final_price = base_price
                            logger.debug(f"Calculated base price from age group: {final_price}€ (age: {age})")
                            break
                else:
                    logger.warning(f"No ruleset or age_groups found, price will be 0.0")

                logger.debug(f"Final calculated price (before family/role discounts): {final_price}€")
        
                # Teilnehmer erstellen
                new_participant = Participant(
                    event_id=event_id,
                    first_name=participant_data["first_name"],
                    last_name=participant_data["last_name"],
                    birth_date=birth_date,
                    gender=participant_data.get("gender"),
                    email=participant_data.get("email"),
                    phone=participant_data.get("phone"),
                    address=participant_data.get("address"),
                    role_id=role_id,
                    family_id=family_id,
                    calculated_price=final_price,
                    is_active=True
                )
        
                db.add(new_participant)
                imported_count += 1
                logger.info(f"✓ Added participant: {new_participant.first_name} {new_participant.last_name}")
        
            except Exception as e:
                logger.exception(f"✗ Error importing participant {participant_data.get('first_name')} {participant_data.get('last_name')}: {e}")
                skipped_count += 1
                continue
        
        # Commit aller Änderungen
        logger.info(f"Committing changes: {imported_count} participants to import")
        db.commit()
        logger.info(f"✓ Import completed: {imported_count} imported, {skipped_count} skipped")

        # Preise für alle importierten Teilnehmer neu berechnen (für Familienrabatte)
        logger.info("Recalculating prices with family and role discounts...")
        all_participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True
        ).all()

        recalculated_count = 0
        for participant in all_participants:
            try:
                # Preis mit PriceCalculator neu berechnen (inkl. Familienrabatte)
                new_price = PriceCalculator.calculate_price_from_db(
                    db=db,
                    event_id=event_id,
                    role_id=participant.role_id,
                    birth_date=participant.birth_date,
                    family_id=participant.family_id
                )
                participant.calculated_price = new_price
                recalculated_count += 1
            except Exception as e:
                logger.warning(f"Could not recalculate price for {participant.full_name}: {e}")

        db.commit()
        logger.info(f"✓ Recalculated prices for {recalculated_count} participants")

        # Erfolgs-Nachricht
        message = f"{imported_count} Teilnehmer erfolgreich importiert"
        if skipped_count > 0:
            message += f" ({skipped_count} übersprungen)"

        flash(request, message, "success")
        return RedirectResponse(url="/participants", status_code=303)

    except json.JSONDecodeError as e:
        logger.exception(f"Invalid JSON in import data: {e}")
        flash(request, "Ungültige Import-Daten", "error")
        return RedirectResponse(url="/participants/import", status_code=303)

    except Exception as e:
        db.rollback()
        logger.exception(f"Error during import: {e}")
        flash(request, f"Fehler beim Import: {str(e)}", "error")
        return RedirectResponse(url="/participants/import", status_code=303)


@router.get("/export")
async def export_participants_excel(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Exportiert alle Teilnehmer als Excel-Datei, gruppiert nach Familien"""
    try:
        # Event laden
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event nicht gefunden")

        # Alle aktiven Teilnehmer laden mit Eager Loading
        all_participants = db.query(Participant).options(
            joinedload(Participant.role),
            joinedload(Participant.family),
            joinedload(Participant.payments)
        ).filter(
            Participant.event_id == event_id,
            Participant.is_active == True
        ).order_by(Participant.last_name, Participant.first_name).all()

        # Workbook erstellen mit ExcelService
        wb, ws = ExcelService.create_workbook("Teilnehmerliste")

        # Header-Zeile
        headers = [
            "Nachname", "Vorname", "Geburtsdatum", "Alter", "Geschlecht",
            "Rolle", "Familie", "E-Mail", "Telefon", "Preis (€)",
            "Bezahlt (€)", "Offen (€)", "Adresse"
        ]

        # Spaltenbreiten
        column_widths = {
            1: 15, 2: 15, 3: 15, 4: 8, 5: 12,
            6: 15, 7: 20, 8: 25, 9: 15, 10: 12,
            11: 12, 12: 12, 13: 30
        }

        # Header-Formatierung mit ExcelService
        ExcelService.apply_header_row(ws, headers, column_widths)

        # Familien gruppieren
        families = db.query(Family).filter(Family.event_id == event_id).order_by(Family.name).all()

        row_num = 2
        family_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Zuerst: Familien mit ihren Mitgliedern
        for family in families:
            family_participants = [p for p in all_participants if p.family_id == family.id]

            if not family_participants:
                continue

            # Familie-Header-Zeile
            ws.cell(row=row_num, column=1, value=f"Familie {family.name}").font = Font(bold=True, size=11)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = family_fill
            row_num += 1

            # Familienmitglieder
            for participant in sorted(family_participants, key=lambda p: p.birth_date):
                _write_participant_row(ws, row_num, participant, event, db)
                row_num += 1

        # Dann: Einzelpersonen ohne Familie
        individual_participants = [p for p in all_participants if p.family_id is None]

        if individual_participants:
            # Leerzeile
            row_num += 1

            # Einzelpersonen-Header
            ws.cell(row=row_num, column=1, value="Einzelpersonen").font = Font(bold=True, size=11)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = family_fill
            row_num += 1

            for participant in individual_participants:
                _write_participant_row(ws, row_num, participant, event, db)
                row_num += 1

        # Summarium am Ende
        row_num += 1
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        summary_font = Font(bold=True)

        ws.cell(row=row_num, column=1, value="GESAMT").font = summary_font
        ws.cell(row=row_num, column=1).fill = summary_fill

        # Gesamtanzahl
        ws.cell(row=row_num, column=7, value=f"{len(all_participants)} Teilnehmer").font = summary_font
        ws.cell(row=row_num, column=7).fill = summary_fill

        # Gesamtpreise
        # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
        total_price = float(sum((p.final_price for p in all_participants), 0))
        total_paid = float(sum((sum((pay.amount for pay in p.payments), 0) for p in all_participants), 0))
        total_outstanding = total_price - total_paid

        ws.cell(row=row_num, column=10, value=total_price).font = summary_font
        ws.cell(row=row_num, column=10).fill = summary_fill
        ws.cell(row=row_num, column=10).number_format = '#,##0.00'

        ws.cell(row=row_num, column=11, value=total_paid).font = summary_font
        ws.cell(row=row_num, column=11).fill = summary_fill
        ws.cell(row=row_num, column=11).number_format = '#,##0.00'

        ws.cell(row=row_num, column=12, value=total_outstanding).font = summary_font
        ws.cell(row=row_num, column=12).fill = summary_fill
        ws.cell(row=row_num, column=12).number_format = '#,##0.00'

        # Excel in BytesIO speichern
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Dateiname mit Event-Name und Datum
        filename = f"Teilnehmerliste_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        logger.exception(f"Error exporting participants: {e}")
        raise HTTPException(status_code=500, detail=f"Fehler beim Export: {str(e)}")


def _write_participant_row(ws: Worksheet, row_num: int, participant: Participant, event: Event, db: Session) -> None:
    """
    Hilfsfunktion zum Schreiben einer Teilnehmer-Zeile in ein Excel-Worksheet.

    Args:
        ws: Worksheet-Objekt
        row_num: Zeilennummer
        participant: Teilnehmer-Objekt
        event: Event-Objekt
        db: Datenbank-Session für Familienzahlungs-Berechnung
    """
    # Alter berechnen
    age = event.start_date.year - participant.birth_date.year
    if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
        age -= 1

    # Zahlungen berechnen
    # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
    direct_payments = float(sum((payment.amount for payment in participant.payments), 0))

    # Anteilige Familienzahlungen berechnen
    family_payment_share = 0.0
    if participant.family_id:
        # Lade Familie mit Zahlungen und Teilnehmern
        family = db.query(Family).filter(Family.id == participant.family_id).options(
            joinedload(Family.payments),
            joinedload(Family.participants).joinedload(Participant.payments)
        ).first()

        if family:
            # Gesamtzahlungen der Familie
            family_total_payments = float(sum((payment.amount for payment in family.payments), 0))

            # Berechne für jedes Familienmitglied die direkten Zahlungen und offenen Beträge
            family_members_outstanding = []
            total_outstanding = 0.0

            for member in family.participants:
                member_direct_payments = float(sum((payment.amount for payment in member.payments), 0))
                member_outstanding = max(0.0, float(member.final_price) - member_direct_payments)
                family_members_outstanding.append({
                    'participant_id': member.id,
                    'outstanding': member_outstanding
                })
                total_outstanding += member_outstanding

            # Verteile die Familienzahlung proportional auf die offenen Beträge
            if total_outstanding > 0:
                # Finde den aktuellen Teilnehmer in der Liste
                for member_data in family_members_outstanding:
                    if member_data['participant_id'] == participant.id:
                        # Anteilige Verteilung basierend auf offenen Beträgen
                        family_payment_share = (member_data['outstanding'] / total_outstanding) * family_total_payments
                        break
            elif family_total_payments > 0:
                # Wenn alle Beträge bezahlt sind, aber noch Familienzahlungen da sind,
                # verteile proportional nach Sollpreis
                family_total_price = float(sum((p.final_price for p in family.participants), 0))
                if family_total_price > 0:
                    family_payment_share = (float(participant.final_price) / family_total_price) * family_total_payments

    total_paid = direct_payments + family_payment_share
    outstanding = float(participant.final_price) - total_paid

    # Daten in Zellen schreiben
    ws.cell(row=row_num, column=1, value=participant.last_name)
    ws.cell(row=row_num, column=2, value=participant.first_name)
    ws.cell(row=row_num, column=3, value=participant.birth_date.strftime("%d.%m.%Y"))
    ws.cell(row=row_num, column=4, value=age)
    ws.cell(row=row_num, column=5, value=participant.gender or "")
    ws.cell(row=row_num, column=6, value=participant.role.display_name if participant.role else "")
    ws.cell(row=row_num, column=7, value=participant.family.name if participant.family else "")
    ws.cell(row=row_num, column=8, value=participant.email or "")
    ws.cell(row=row_num, column=9, value=participant.phone or "")

    # Preise mit Formatierung
    price_cell = ws.cell(row=row_num, column=10, value=participant.final_price)
    price_cell.number_format = '#,##0.00'

    paid_cell = ws.cell(row=row_num, column=11, value=total_paid)
    paid_cell.number_format = '#,##0.00'

    outstanding_cell = ws.cell(row=row_num, column=12, value=outstanding)
    outstanding_cell.number_format = '#,##0.00'

    # Offener Betrag rot markieren wenn > 0
    if outstanding > 0:
        outstanding_cell.font = Font(color="FF0000", bold=True)

    ws.cell(row=row_num, column=13, value=participant.address or "")


# ===== PARAMETRISIERTE ROUTES (müssen NACH den spezifischen Routes kommen) =====

@router.get("/{participant_id}", response_class=HTMLResponse)
async def view_participant(
    request: Request,
    participant_id: int,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Detailansicht eines Teilnehmers"""
    participant = db.query(Participant).options(
        joinedload(Participant.role),
        joinedload(Participant.family),
        joinedload(Participant.event),
        joinedload(Participant.payments)
    ).filter(
        Participant.id == participant_id,
        Participant.event_id == event_id
    ).first()

    if not participant:
        return RedirectResponse(url="/participants", status_code=303)

    # Zahlungen des Teilnehmers
    # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
    direct_payments = float(sum((payment.amount for payment in participant.payments), 0))

    # Anteilige Familienzahlungen berechnen
    family_payment_share = 0.0
    if participant.family_id:
        # Lade Familie mit Zahlungen und Teilnehmern
        family = db.query(Family).filter(Family.id == participant.family_id).options(
            joinedload(Family.payments),
            joinedload(Family.participants).joinedload(Participant.payments)
        ).first()

        if family:
            # Gesamtzahlungen der Familie
            family_total_payments = float(sum((payment.amount for payment in family.payments), 0))

            # Berechne für jedes Familienmitglied die direkten Zahlungen und offenen Beträge
            family_members_outstanding = []
            total_outstanding = 0.0

            for member in family.participants:
                member_direct_payments = float(sum((payment.amount for payment in member.payments), 0))
                member_outstanding = max(0.0, float(member.final_price) - member_direct_payments)
                family_members_outstanding.append({
                    'participant_id': member.id,
                    'outstanding': member_outstanding
                })
                total_outstanding += member_outstanding

            # Verteile die Familienzahlung proportional auf die offenen Beträge
            if total_outstanding > 0:
                # Finde den aktuellen Teilnehmer in der Liste
                for member_data in family_members_outstanding:
                    if member_data['participant_id'] == participant.id:
                        # Anteilige Verteilung basierend auf offenen Beträgen
                        family_payment_share = (member_data['outstanding'] / total_outstanding) * family_total_payments
                        break
            elif family_total_payments > 0:
                # Wenn alle Beträge bezahlt sind, aber noch Familienzahlungen da sind,
                # verteile proportional nach Sollpreis
                family_total_price = float(sum((p.final_price for p in family.participants), 0))
                if family_total_price > 0:
                    family_payment_share = (float(participant.final_price) / family_total_price) * family_total_payments

    total_paid = direct_payments + family_payment_share
    outstanding = float(participant.final_price) - total_paid

    return templates.TemplateResponse(
        "participants/detail.html",
        {
            "request": request,
            "title": f"{participant.full_name}",
            "participant": participant,
            "total_paid": total_paid,
            "outstanding": outstanding
        }
    )


@router.get("/{participant_id}/edit", response_class=HTMLResponse)
async def edit_participant_form(
    request: Request,
    participant_id: int,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Formular zum Bearbeiten eines Teilnehmers"""
    participant = db.query(Participant).options(
        joinedload(Participant.role),
        joinedload(Participant.family),
        joinedload(Participant.event)
    ).filter(
        Participant.id == participant_id,
        Participant.event_id == event_id
    ).first()

    if not participant:
        return RedirectResponse(url="/participants", status_code=303)

    roles = db.query(Role).filter(Role.is_active == True, Role.event_id == event_id).all()
    event = db.query(Event).filter(Event.id == event_id).first()
    families = db.query(Family).filter(Family.event_id == event_id).order_by(Family.name).all()

    return templates.TemplateResponse(
        "participants/edit.html",
        {
            "request": request,
            "title": f"{participant.full_name} bearbeiten",
            "participant": participant,
            "roles": roles,
            "event": event,
            "families": families
        }
    )


@router.post("/{participant_id}/edit", response_class=HTMLResponse)
async def update_participant(
    request: Request,
    participant_id: int,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    first_name: str = Form(...),
    last_name: str = Form(...),
    birth_date: str = Form(...),
    gender: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    bildung_teilhabe_id: Optional[str] = Form(None),
    allergies: Optional[str] = Form(None),
    medical_notes: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    discount_percent: str = Form("0.0"),  # Als String empfangen
    discount_reason: Optional[str] = Form(None),
    manual_price_override: str = Form(""),  # Als String empfangen
    role_id: Optional[str] = Form(None),  # Als String empfangen
    family_id: Optional[str] = Form(None)  # Als String empfangen
):
    """Aktualisiert einen Teilnehmer"""
    participant = db.query(Participant).options(
        joinedload(Participant.role),
        joinedload(Participant.family)
    ).filter(
        Participant.id == participant_id,
        Participant.event_id == event_id
    ).first()

    if not participant:
        return RedirectResponse(url="/participants", status_code=303)

    try:
        # Konvertiere leere Strings zu None oder passenden Werten
        email_val = None if not email or email.strip() == "" else email
        manual_price_override_val = None if not manual_price_override or manual_price_override.strip() == "" else float(manual_price_override)
        discount_percent_val = 0.0 if not discount_percent or discount_percent.strip() == "" else float(discount_percent)
        role_id_val = None if not role_id or role_id.strip() == "" else int(role_id)
        family_id_val = None if not family_id or family_id.strip() == "" else int(family_id)

        # Pydantic-Validierung
        participant_data = ParticipantUpdateSchema(
            first_name=first_name,
            last_name=last_name,
            birth_date=birth_date,
            gender=gender,
            email=email_val,
            phone=phone,
            address=address,
            bildung_teilhabe_id=bildung_teilhabe_id,
            allergies=allergies,
            medical_notes=medical_notes,
            notes=notes,
            discount_percent=discount_percent_val,
            discount_reason=discount_reason,
            manual_price_override=manual_price_override_val,
            role_id=role_id_val,
            family_id=family_id_val
        )

        # Preis neu berechnen (wenn sich relevante Daten geändert haben)
        # birth_date ist bereits ein date-Objekt nach Pydantic-Validierung
        calculated_price = _calculate_price_for_participant(
            db=db,
            event_id=event_id,
            role_id=participant_data.role_id,
            birth_date=participant_data.birth_date,
            family_id=participant_data.family_id
        )

        # Teilnehmer aktualisieren (birth_date ist bereits ein date-Objekt)
        participant.first_name = participant_data.first_name
        participant.last_name = participant_data.last_name
        participant.birth_date = participant_data.birth_date
        participant.gender = participant_data.gender
        participant.email = participant_data.email
        participant.phone = participant_data.phone
        participant.address = participant_data.address
        participant.bildung_teilhabe_id = participant_data.bildung_teilhabe_id
        participant.allergies = participant_data.allergies
        participant.medical_notes = participant_data.medical_notes
        participant.notes = participant_data.notes
        participant.discount_percent = participant_data.discount_percent
        participant.discount_reason = participant_data.discount_reason
        participant.manual_price_override = participant_data.manual_price_override
        # event_id bleibt unverändert (Sicherheit!)

        # Speichere alte role_id für Rollenüberschreitungs-Prüfung
        old_role_id = participant.role_id

        participant.role_id = participant_data.role_id
        participant.family_id = participant_data.family_id
        participant.calculated_price = calculated_price

        db.commit()

        # Prüfe Rollenüberschreitung für beide Rollen (alte und neue), falls geändert
        if old_role_id != participant.role_id:
            # Alte Rolle: Prüfe ob Überschreitung behoben wurde
            if old_role_id:
                _check_and_update_role_count_task(db, event_id, old_role_id)
            # Neue Rolle: Prüfe ob neue Überschreitung entstanden ist
            if participant.role_id:
                _check_and_update_role_count_task(db, event_id, participant.role_id)
            db.commit()  # Speichere Task-Änderungen
        else:
            # Rolle hat sich nicht geändert, aber prüfe trotzdem (z.B. für manuelle Korrekturen)
            _check_and_update_role_count_task(db, event_id, participant.role_id)
            db.commit()  # Speichere Task-Änderungen

        flash(request, f"Teilnehmer {participant.full_name} wurde erfolgreich aktualisiert", "success")
        return RedirectResponse(url=f"/participants/{participant_id}", status_code=303)

    except ValidationError as e:
        # Pydantic-Validierungsfehler
        logger.warning(f"Validation error updating participant: {e}", exc_info=True)
        first_error = e.errors()[0]
        field_name = first_error['loc'][0] if first_error['loc'] else 'Unbekannt'
        error_msg = first_error['msg']
        flash(request, f"Validierungsfehler ({field_name}): {error_msg}", "error")
        return RedirectResponse(url=f"/participants/{participant_id}/edit?error=validation", status_code=303)

    except IntegrityError as e:
        db.rollback()
        logger.exception(f"Database integrity error updating participant: {e}")
        flash(request, "Teilnehmer konnte nicht aktualisiert werden (Datenbankfehler)", "error")
        return RedirectResponse(url=f"/participants/{participant_id}/edit?error=db_integrity", status_code=303)

    except DataError as e:
        db.rollback()
        logger.exception(f"Invalid data updating participant: {e}")
        flash(request, "Ungültige Daten eingegeben", "error")
        return RedirectResponse(url=f"/participants/{participant_id}/edit?error=invalid_data", status_code=303)

    except Exception as e:
        return handle_db_exception(e, f"/participants/{participant_id}/edit", "Updating participant", db, request)


@router.post("/{participant_id}/delete")
async def delete_participant(
    request: Request,
    participant_id: int,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Löscht einen Teilnehmer"""
    participant = db.query(Participant).filter(
        Participant.id == participant_id,
        Participant.event_id == event_id
    ).first()

    if not participant:
        flash(request, "Teilnehmer nicht gefunden", "error")
        return RedirectResponse(url="/participants", status_code=303)

    try:
        participant_name = participant.full_name
        participant_role_id = participant.role_id  # Speichere für Rollenüberschreitungs-Prüfung

        # Soft-Delete: Statt db.delete() markieren wir als gelöscht
        participant.is_active = False
        participant.deleted_at = utcnow()
        db.commit()

        # Prüfe Rollenüberschreitung (möglicherweise wurde sie durch Löschen behoben)
        _check_and_update_role_count_task(db, event_id, participant_role_id)
        db.commit()  # Speichere Task-Änderungen

        logger.info(f"Participant soft-deleted: {participant_name} (ID: {participant_id})")
        flash(request, f"Teilnehmer {participant_name} wurde erfolgreich gelöscht", "success")
        return RedirectResponse(url="/participants", status_code=303)

    except Exception as e:
        db.rollback()
        logger.exception(f"Error deleting participant: {e}")
        flash(request, "Fehler beim Löschen des Teilnehmers", "error")
        return RedirectResponse(url="/participants", status_code=303)

@router.get("/{participant_id}/payment-qr", response_class=Response)
async def generate_payment_qr_code(
    participant_id: int,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Generiert einen QR-Code für die Zahlung (SEPA EPC QR-Code)"""
    # Teilnehmer laden
    participant = db.query(Participant).filter(
        Participant.id == participant_id,
        Participant.event_id == event_id
    ).first()

    if not participant:
        raise HTTPException(status_code=404, detail="Teilnehmer nicht gefunden")

    # Einstellungen laden
    setting = db.query(Setting).filter(Setting.event_id == event_id).first()

    if not setting or not setting.bank_iban:
        raise HTTPException(status_code=400, detail="Bank-Daten nicht konfiguriert")

    # Berechne offenen Betrag
    # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
    total_paid = float(sum((payment.amount for payment in participant.payments), 0))
    outstanding = float(participant.final_price) - total_paid

    if outstanding <= 0:
        outstanding = participant.final_price  # Zeige Gesamtpreis wenn bereits bezahlt

    # Verwendungszweck erstellen
    invoice_number = f"TN-{participant.id:06d}"
    purpose = f"{participant.event.name} - {participant.full_name} - Rechnungsnr: {invoice_number}"

    # QR-Code generieren
    qr_code_data = QRCodeService.generate_sepa_qr_code(
        recipient_name=setting.organization_name or "Freizeit-Organisation",
        iban=setting.bank_iban,
        amount=outstanding,
        purpose=purpose,
        bic=setting.bank_bic
    )

    return Response(
        content=qr_code_data,
        media_type="image/png",
        headers={
            "Content-Disposition": f"inline; filename=payment_qr_{participant_id}.png"
        }
    )
